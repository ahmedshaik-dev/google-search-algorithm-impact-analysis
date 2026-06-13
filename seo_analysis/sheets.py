import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet


ENTITY_COLUMNS = [
    'Entity',
    'Clicks_Before', 'Clicks_After', 'Clicks_Change', 'Clicks_Pct',
    'Impressions_Before', 'Impressions_After', 'Impressions_Change', 'Impressions_Pct',
    'CTR_Before', 'CTR_After', 'CTR_Change', 'CTR_Pct',
    'Position_Before', 'Position_After', 'Pos_Change', 'Pos_Pct',
    'Impact_Score', 'Status', 'Category',
]


def _write_title(ws: Worksheet, title: str, row: int = 1) -> None:
    ws.cell(row=row, column=1, value=title)


def _write_headers(ws: Worksheet, headers: list, row: int) -> None:
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=row, column=col_idx, value=header)


def _write_df(ws: Worksheet, df: pd.DataFrame, start_row: int) -> None:
    for row_idx, row_data in enumerate(df.itertuples(index=False), start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def build_summary_sheet(ws: Worksheet, data: dict, client_name: str, date_labels: dict) -> None:
    """Write Summary_Dashboard with KPIs from all sheets."""
    queries_df = data.get('queries')

    ws.cell(row=1, column=1, value=f'SEO Performance Report — {client_name}')
    ws.cell(row=2, column=1, value=f'Period: {date_labels["before"]}  →  {date_labels["after"]}')

    if queries_df is None or queries_df.empty:
        ws.cell(row=4, column=1, value='No query data available.')
        return

    total_kw = len(queries_df)
    total_clicks_b = int(queries_df['Clicks_Before'].sum())
    total_clicks_a = int(queries_df['Clicks_After'].sum())
    total_imp_b = int(queries_df['Impressions_Before'].sum())
    total_imp_a = int(queries_df['Impressions_After'].sum())
    avg_pos_b = round(float(queries_df['Position_Before'].mean()), 1)
    avg_pos_a = round(float(queries_df['Position_After'].mean()), 1)
    win_rate = round(len(queries_df[queries_df['Impact_Score'] > 0]) / total_kw * 100, 1)

    pos_impr = (avg_pos_b - avg_pos_a) / avg_pos_b * 100 if avg_pos_b > 0 else 0
    health = round(min(100, max(0, win_rate * 0.6 + max(0, pos_impr) * 0.4)), 1)

    rows = [
        ('Metric', 'Before', 'After', 'Change'),
        ('Total Keywords', total_kw, total_kw, '—'),
        ('Total Clicks', total_clicks_b, total_clicks_a,
         f'{total_clicks_a - total_clicks_b:+d}'),
        ('Total Impressions', total_imp_b, total_imp_a,
         f'{total_imp_a - total_imp_b:+d}'),
        ('Avg Position', avg_pos_b, avg_pos_a,
         f'{avg_pos_a - avg_pos_b:+.1f}'),
        ('', '', '', ''),
        ('Win Rate', f'{win_rate}%', '', '% keywords with positive score'),
        ('Portfolio Health Score', f'{health}/100', '', ''),
    ]

    for r_idx, row_vals in enumerate(rows, start=4):
        for c_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # Status breakdown
    ws.cell(row=13, column=1, value='Keyword Status Breakdown')
    ws.cell(row=14, column=1, value='Status')
    ws.cell(row=14, column=2, value='Count')
    ws.cell(row=14, column=3, value='% of Total')
    for s_idx, status in enumerate(['GAINED', 'IMPROVED', 'FLAT', 'DECLINED', 'LOST'], start=15):
        count = len(queries_df[queries_df['Status'] == status])
        ws.cell(row=s_idx, column=1, value=status)
        ws.cell(row=s_idx, column=2, value=count)
        ws.cell(row=s_idx, column=3, value=f'{count / total_kw * 100:.1f}%')


def build_entity_sheet(ws: Worksheet, df: pd.DataFrame, title: str) -> None:
    """
    Write a standard entity analysis sheet (Queries, Pages, Countries).
    Sorted by Impact_Score descending.
    """
    _write_title(ws, title, row=1)

    cols = [c for c in ENTITY_COLUMNS if c in df.columns]
    _write_headers(ws, cols, row=2)

    sorted_df = df.sort_values('Impact_Score', ascending=False)[cols]
    _write_df(ws, sorted_df, start_row=3)


def build_devices_sheet(ws: Worksheet, df: pd.DataFrame) -> None:
    """Write Devices_Analysis sheet — traffic split across Mobile/Desktop/Tablet."""
    _write_title(ws, 'Devices Analysis', row=1)
    cols = [c for c in ENTITY_COLUMNS if c in df.columns]
    _write_headers(ws, cols, row=2)
    _write_df(ws, df[cols], start_row=3)


def build_search_appearance_sheet(ws: Worksheet, df: pd.DataFrame) -> None:
    """Write Search_Appearance_Analysis — dynamic rich result types per client."""
    _write_title(ws, 'Search Appearance Analysis', row=1)
    cols = [c for c in ENTITY_COLUMNS if c in df.columns]
    _write_headers(ws, cols, row=2)
    _write_df(ws, df[cols], start_row=3)


def build_algo_impact_sheet(ws: Worksheet, df: pd.DataFrame,
                             summary: dict, update_date: str) -> None:
    """Write Algorithm_Impact sheet with summary stats and flagged keywords."""
    _write_title(ws, f'Algorithm Impact Analysis — Update date: {update_date}', row=1)

    summary_rows = [
        ('Total Keywords', summary['total_keywords']),
        ('Algo-Affected Keywords', f"{summary['affected_count']} ({summary['affected_pct']}%)"),
        ('Gained (impressions up)', summary['positive_count']),
        ('Lost (impressions down)', summary['negative_count']),
        ('Total Impression Impact', summary['total_impression_impact']),
        ('Severity', summary['severity']),
    ]
    for r_idx, (label, value) in enumerate(summary_rows, start=2):
        ws.cell(row=r_idx, column=1, value=label)
        ws.cell(row=r_idx, column=2, value=value)

    affected_df = df[df['Algo_Affected']].copy()
    affected_df = affected_df.sort_values('Impact_Score')

    start_row = 10
    ws.cell(row=start_row, column=1, value='Algo-Affected Keywords (sorted by impact)')
    cols = [c for c in ENTITY_COLUMNS if c in affected_df.columns]
    _write_headers(ws, cols, row=start_row + 1)
    _write_df(ws, affected_df[cols], start_row=start_row + 2)
