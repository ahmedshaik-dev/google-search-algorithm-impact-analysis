from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Font, Alignment
from typing import List, Dict


COLOR_MAP = {
    'STRONG_POSITIVE': 'C6EFCE',  # green
    'POSITIVE': 'DDEBF7',         # light blue
    'WATCH': 'FFEB9C',            # yellow
    'NEGATIVE': 'FFC7CE',         # red
}

STATUS_COLOR_MAP = {
    'GAINED': 'C6EFCE',
    'IMPROVED': 'DDEBF7',
    'FLAT': 'F2F2F2',
    'DECLINED': 'FFC7CE',
    'LOST': 'FF0000',
}

HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
TITLE_FONT = Font(bold=True, size=13, color='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF')


def _apply_title_style(ws: Worksheet, row: int = 1) -> None:
    cell = ws.cell(row=row, column=1)
    cell.font = TITLE_FONT


def _apply_header_row(ws: Worksheet, row: int) -> None:
    for cell in ws[row]:
        if cell.value:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center')


def _apply_category_colors(ws: Worksheet, category_col: int, data_start_row: int) -> None:
    for row in ws.iter_rows(min_row=data_start_row):
        cat_cell = row[category_col - 1]
        color = COLOR_MAP.get(str(cat_cell.value), None)
        if color:
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cat_cell.fill = fill


def _set_column_widths(ws: Worksheet) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, val_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)


def format_workbook(wb: Workbook, sheet_configs: List[Dict]) -> None:
    """
    Apply formatting to all sheets.

    sheet_configs: list of dicts with keys:
        sheet: Worksheet
        header_row: int (row index of column headers)
        category_col: int (1-based column index of Category values)
        data_start_row: int (first row with data)
    """
    for config in sheet_configs:
        ws = config['sheet']
        header_row = config.get('header_row', 1)
        category_col = config.get('category_col')
        data_start_row = config.get('data_start_row', header_row + 1)

        _apply_title_style(ws, row=1)
        _apply_header_row(ws, row=header_row)

        if category_col:
            _apply_category_colors(ws, category_col, data_start_row)

        _set_column_widths(ws)

        if ws.max_row >= header_row:
            ws.freeze_panes = ws.cell(row=data_start_row, column=2)
