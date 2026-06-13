import argparse
import sys
from pathlib import Path
from openpyxl import Workbook

from .reader import load_gsc_export
from .calculator import calculate_metrics
from .classifier import classify_keywords
from .algo_detector import detect_impact, build_algo_summary
from .sheets import (
    build_summary_sheet, build_entity_sheet,
    build_devices_sheet, build_search_appearance_sheet,
    build_algo_impact_sheet, ENTITY_COLUMNS,
)
from .formatter import format_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog='seo_analysis',
        description='Generate SEO analysis report from a GSC comparison export.',
    )
    parser.add_argument('--input', required=True, help='Path to GSC export .xlsx')
    parser.add_argument('--output', required=True, help='Path for output report .xlsx')
    parser.add_argument('--client', default='Client', help='Client name for report headers')
    parser.add_argument('--mode', default='standard', choices=['standard', 'algo_update'],
                        help='Analysis mode')
    parser.add_argument('--update-date', dest='update_date', default=None,
                        help='Algorithm update date (YYYY-MM-DD), required for algo_update mode')
    return parser.parse_args()


def _process_entity_df(df):
    """Apply metrics calculation and classification to any entity DataFrame."""
    if df is None or df.empty:
        return None
    return classify_keywords(calculate_metrics(df))


def generate_report(input_path: str, output_path: str, client: str,
                    mode: str, update_date: str = None) -> None:
    """Load GSC export, compute all metrics, write output workbook."""
    print(f"Loading: {input_path}")
    data = load_gsc_export(input_path)

    print("Computing metrics...")
    data['queries'] = _process_entity_df(data['queries'])
    data['pages'] = _process_entity_df(data['pages'])
    data['countries'] = _process_entity_df(data['countries'])
    data['devices'] = _process_entity_df(data['devices'])
    data['search_appearance'] = _process_entity_df(data['search_appearance'])

    wb = Workbook()
    wb.remove(wb.active)  # Remove default empty sheet
    sheet_configs = []

    # Sheet 1: Summary
    ws_summary = wb.create_sheet('Summary_Dashboard')
    build_summary_sheet(ws_summary, data, client, data['date_labels'])
    sheet_configs.append({'sheet': ws_summary, 'header_row': 4, 'category_col': None,
                          'data_start_row': 5})

    # Helper: find Category column index
    def _cat_col_idx(df):
        if df is None:
            return None
        cols = [c for c in ENTITY_COLUMNS if c in df.columns]
        return cols.index('Category') + 1 if 'Category' in cols else None

    # Sheet 2: Queries
    if data['queries'] is not None:
        ws_q = wb.create_sheet('Queries_Analysis')
        build_entity_sheet(ws_q, data['queries'], 'Queries Analysis')
        sheet_configs.append({'sheet': ws_q, 'header_row': 2,
                               'category_col': _cat_col_idx(data['queries']),
                               'data_start_row': 3})

    # Sheet 3: Pages
    if data['pages'] is not None:
        ws_p = wb.create_sheet('Pages_Analysis')
        build_entity_sheet(ws_p, data['pages'], 'Pages Analysis')
        sheet_configs.append({'sheet': ws_p, 'header_row': 2,
                               'category_col': _cat_col_idx(data['pages']),
                               'data_start_row': 3})

    # Sheet 4: Countries
    if data['countries'] is not None:
        ws_c = wb.create_sheet('Countries_Analysis')
        build_entity_sheet(ws_c, data['countries'], 'Countries Analysis')
        sheet_configs.append({'sheet': ws_c, 'header_row': 2,
                               'category_col': _cat_col_idx(data['countries']),
                               'data_start_row': 3})

    # Sheet 5: Devices
    if data['devices'] is not None:
        ws_d = wb.create_sheet('Devices_Analysis')
        build_devices_sheet(ws_d, data['devices'])
        sheet_configs.append({'sheet': ws_d, 'header_row': 2,
                               'category_col': _cat_col_idx(data['devices']),
                               'data_start_row': 3})

    # Sheet 6: Search Appearance
    if data['search_appearance'] is not None:
        ws_sa = wb.create_sheet('Search_Appearance_Analysis')
        build_search_appearance_sheet(ws_sa, data['search_appearance'])
        sheet_configs.append({'sheet': ws_sa, 'header_row': 2,
                               'category_col': _cat_col_idx(data['search_appearance']),
                               'data_start_row': 3})

    # Sheet 7: Algorithm Impact (conditional)
    if mode == 'algo_update' and data['queries'] is not None:
        queries_with_algo = detect_impact(data['queries'])
        summary = build_algo_summary(queries_with_algo)
        ws_algo = wb.create_sheet('Algorithm_Impact')
        build_algo_impact_sheet(ws_algo, queries_with_algo, summary, update_date)
        sheet_configs.append({'sheet': ws_algo, 'header_row': 11,
                               'category_col': _cat_col_idx(data['queries']),
                               'data_start_row': 12})
        print(f"Algorithm Impact: {summary['severity']} severity, "
              f"{summary['affected_count']} keywords affected ({summary['affected_pct']}%)")

    print("Applying formatting...")
    format_workbook(wb, sheet_configs)

    wb.save(output_path)
    print(f"Report saved: {output_path}")


def main() -> None:
    args = parse_args()

    if not Path(args.input).exists():
        print(f"Error: Input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    if args.mode == 'algo_update' and not args.update_date:
        print("Error: --update-date is required when --mode=algo_update", file=sys.stderr)
        sys.exit(1)

    generate_report(
        input_path=args.input,
        output_path=args.output,
        client=args.client,
        mode=args.mode,
        update_date=args.update_date,
    )


if __name__ == '__main__':
    main()
