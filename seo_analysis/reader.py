import re
import pandas as pd
import numpy as np
from datetime import datetime
from typing import Optional, Dict, Tuple


def parse_column_headers(headers: list) -> dict:
    """
    Parse GSC export column headers into metric mapping.

    Input: ['Top queries', '10/05/2026 - 10/05/2026 Clicks', '11/04/2026 - 11/04/2026 Clicks', ...]
    Output: {
        'entity_col': 'Top queries',
        'before_date': '11/04/2026 - 11/04/2026',
        'after_date': '10/05/2026 - 10/05/2026',
        'Clicks': ('11/04/2026 - 11/04/2026 Clicks', '10/05/2026 - 10/05/2026 Clicks'),
        'Impressions': (...), 'CTR': (...), 'Position': (...),
    }
    """
    entity_col = headers[0]
    date_metric_map: Dict[str, Dict[str, str]] = {}

    for h in headers[1:]:
        if h is None:
            continue
        m = re.match(r'(\d{2}/\d{2}/\d{4} - \d{2}/\d{2}/\d{4})\s+(.+)', str(h))
        if m:
            date_range = m.group(1)
            metric = m.group(2).strip()
            if date_range not in date_metric_map:
                date_metric_map[date_range] = {}
            date_metric_map[date_range][metric] = h

    if len(date_metric_map) != 2:
        return {'entity_col': entity_col}

    dates = list(date_metric_map.keys())
    d1 = datetime.strptime(dates[0].split(' - ')[0].strip(), '%d/%m/%Y')
    d2 = datetime.strptime(dates[1].split(' - ')[0].strip(), '%d/%m/%Y')

    before_date = dates[0] if d1 < d2 else dates[1]
    after_date = dates[1] if d1 < d2 else dates[0]

    result: dict = {
        'entity_col': entity_col,
        'before_date': before_date,
        'after_date': after_date,
    }

    for metric in ['Clicks', 'Impressions', 'CTR', 'Position']:
        before_col = date_metric_map[before_date].get(metric)
        after_col = date_metric_map[after_date].get(metric)
        if before_col and after_col:
            result[metric] = (before_col, after_col)

    return result


def _standardize_df(df: pd.DataFrame, col_info: dict) -> pd.DataFrame:
    """Rename date-prefixed columns to Before/After names; auto-fill missing numeric values with 0."""
    entity_col = col_info.get('entity_col', df.columns[0])
    rename_map = {entity_col: 'Entity'}

    for metric in ['Clicks', 'Impressions', 'CTR', 'Position']:
        if metric in col_info:
            before_col, after_col = col_info[metric]
            rename_map[before_col] = f'{metric}_Before'
            rename_map[after_col] = f'{metric}_After'

    df = df.rename(columns=rename_map)
    df = df.dropna(subset=['Entity'])

    numeric_cols = [c for c in df.columns if c != 'Entity']
    missing_count = int(df[numeric_cols].isna().sum().sum())
    total_cells = len(df) * len(numeric_cols)
    if total_cells > 0 and missing_count / total_cells > 0.1:
        pct = missing_count / total_cells * 100
        print(f"Warning: {missing_count} missing values ({pct:.0f}% of data), auto-filling with 0")
    df[numeric_cols] = df[numeric_cols].fillna(0)

    return df.reset_index(drop=True)


def load_gsc_export(filepath: str) -> dict:
    """
    Load all sheets from a GSC comparison export (.xlsx).

    Returns:
        {
            'queries': pd.DataFrame or None,
            'pages': pd.DataFrame or None,
            'countries': pd.DataFrame or None,
            'devices': pd.DataFrame or None,
            'search_appearance': pd.DataFrame or None,
            'date_labels': {'before': str, 'after': str},
        }
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    sheet_map = {
        'Queries': 'queries',
        'Pages': 'pages',
        'Countries': 'countries',
        'Devices': 'devices',
        'Search appearance': 'search_appearance',
    }

    result: dict = {v: None for v in sheet_map.values()}
    result['date_labels'] = {'before': '', 'after': ''}

    for sheet_name, key in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found, skipping.")
            continue

        ws = wb[sheet_name]
        rows = [
            tuple(row)
            for row in ws.iter_rows(values_only=True)
            if any(v is not None for v in row)
        ]

        if not rows:
            continue

        headers = list(rows[0])
        data = rows[1:]
        df = pd.DataFrame(data, columns=headers)

        col_info = parse_column_headers(headers)

        if key == 'queries' and 'before_date' in col_info:
            result['date_labels']['before'] = col_info['before_date']
            result['date_labels']['after'] = col_info['after_date']

        result[key] = _standardize_df(df, col_info)

    wb.close()
    return result
